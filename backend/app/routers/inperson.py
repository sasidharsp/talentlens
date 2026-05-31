import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models
from app.auth import require_any_staff

router = APIRouter(prefix="/api/inperson", tags=["inperson"])


@router.get("/tags")
def get_tags(db: Session = Depends(get_db),
             current_user=Depends(require_any_staff)):
    """All unique tags in use, with question count each."""
    rows = db.query(
        models.InPersonQuestion.tag,
        models.func.count(models.InPersonQuestion.id).label("count")
    ).group_by(models.InPersonQuestion.tag)\
     .order_by(models.InPersonQuestion.tag).all()
    return [{"tag": r.tag, "count": r.count} for r in rows]


@router.get("/questions")
def get_questions(
    tag: str = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_any_staff),
):
    """List questions, optionally filtered by tag."""
    q = db.query(models.InPersonQuestion)
    if tag:
        q = q.filter(models.InPersonQuestion.tag == tag)
    items = q.order_by(models.InPersonQuestion.tag,
                       models.InPersonQuestion.id).all()
    return [
        {
            "id": i.id,
            "question": i.question,
            "answer": i.answer,
            "tag": i.tag,
            "created_at": i.created_at,
        }
        for i in items
    ]


@router.get("/stats")
def get_stats(db: Session = Depends(get_db),
              current_user=Depends(require_any_staff)):
    total_q   = db.query(models.InPersonQuestion).count()
    total_tags = db.query(models.InPersonQuestion.tag)\
                   .distinct().count()
    return {"total_questions": total_q, "total_tags": total_tags}


@router.post("/questions")
def add_question(
    payload: dict,
    db: Session = Depends(get_db),
    current_user=Depends(require_any_staff),
):
    """Add a new in-person interview question."""
    question = (payload.get("question") or "").strip()
    answer   = (payload.get("answer")   or "").strip()
    tag      = (payload.get("tag")      or "").strip()

    if not question: raise HTTPException(400, "Question is required.")
    if not answer:   raise HTTPException(400, "Answer is required.")
    if not tag:      raise HTTPException(400, "Tag is required.")

    q = models.InPersonQuestion(
        question=question,
        answer=answer,
        tag=tag,
        created_by=current_user.id,
    )
    db.add(q); db.commit(); db.refresh(q)
    return {"id": q.id, "question": q.question, "answer": q.answer, "tag": q.tag}


@router.get("/template")
def download_template(current_user=Depends(require_any_staff)):
    """Download Excel template for bulk importing in-person interview questions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    # Styles
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill("solid", fgColor="312E81")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    SMP_FILL  = PatternFill("solid", fgColor="EEF2FF")
    border    = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = ["tag", "question", "answer"]
    col_widths = [20, 60, 70]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, i, h)
        cell.font    = HDR_FONT
        cell.fill    = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border  = border
        ws.column_dimensions[cell.column_letter].width = w

    # Sample rows
    samples = [
        ["Leadership", "Tell me about a time you led a team through a difficult change.", "Look for: clear ownership, empathy, structured approach, measurable outcome."],
        ["Technical",  "How would you design a high-availability system for a banking application?", "Key points: redundancy, failover, DR strategy, RPO/RTO, compliance."],
        ["Behavioural","Describe a situation where you had to influence without authority.", "Look for: stakeholder mapping, communication style, outcome achieved."],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.fill      = SMP_FILL
            cell.font      = Font(size=10, italic=True, color="374151")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[r].height = 40

    ws.row_dimensions[1].height = 22

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 90
    ws2.cell(1, 1, "TalentLens — In-person Interview Question Import").font = Font(bold=True, size=14, color="312E81")
    for i, line in enumerate([
        "• Column 'tag'      — Topic tag (e.g. Leadership, Technical, Behavioural). Required.",
        "• Column 'question' — The interview question to ask. Required.",
        "• Column 'answer'   — Expected answer or key points the interviewer should look for. Required.",
        "• Delete the sample rows (rows 2–4) before uploading your real questions.",
        "• Do NOT rename the column headers.",
        "• Save as .xlsx before uploading.",
    ], 3):
        c = ws2.cell(i, 1, line)
        c.font = Font(size=11, color="1F2937")
        c.alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inperson_questions_template.xlsx"},
    )


@router.post("/import")
async def import_questions(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_any_staff),
):
    """Bulk import in-person interview questions from Excel or CSV."""
    filename = (file.filename or "").lower()
    content  = await file.read()

    if not content:
        raise HTTPException(400, "Uploaded file is empty.")

    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(400, "Please upload a .xlsx or .csv file.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not read file: {e}")

    if df.empty:
        raise HTTPException(400, "File has no data rows.")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    df.dropna(how="all", inplace=True)

    created, errors = 0, []
    for idx, row in df.iterrows():
        rv     = row.where(pd.notna(row), None).to_dict()
        _s     = lambda k: str(rv[k]).strip() if rv.get(k) not in (None, "") else ""
        tag    = _s("tag")
        question = _s("question")
        answer = _s("answer")

        if not tag or not question or not answer:
            errors.append({"row": idx + 2, "error": "tag, question and answer are all required."})
            continue

        db.add(models.InPersonQuestion(
            tag=tag, question=question, answer=answer,
            created_by=current_user.id,
        ))
        created += 1

    db.commit()
    return {
        "imported": created,
        "errors":   errors,
        "message":  f"{created} question(s) imported successfully.",
    }
