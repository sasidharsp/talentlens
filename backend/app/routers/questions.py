"""
Question Bank Management Routes
Supports CRUD, bulk import (CSV/Excel), bulk export, template download, purge.
"""
import io
import json
from typing import Optional, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app import models, schemas
from app.auth import require_admin

router = APIRouter(prefix="/api/admin/questions", tags=["questions"])


# ══════════════════════════════════════════════════════
#  TEMPLATE HELPERS
# ══════════════════════════════════════════════════════
def _styled_workbook(headers: list, sample_rows: list, notes: list = None, sheet_name="Questions"):
    """Build a styled Excel workbook with headers, sample data, and an instructions sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Colours
    HDR_FILL  = PatternFill("solid", fgColor="312E81")   # deep indigo
    REQ_FILL  = PatternFill("solid", fgColor="4F46E5")   # indigo (required cols)
    OPT_FILL  = PatternFill("solid", fgColor="6D7280")   # gray (optional cols)
    SAMPLE_FILL = PatternFill("solid", fgColor="EEF2FF") # light indigo rows
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Column meta: (header_name, is_required, width)
    for col_idx, (header, required, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill  = REQ_FILL if required else OPT_FILL
        cell.font  = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 32

    # Sample rows
    for row_idx, row_data in enumerate(sample_rows, 2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = SAMPLE_FILL
            cell.font = Font(size=10, italic=True, color="374151")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 90
    ws2.cell(1, 1, "TalentLens — Question Import Instructions").font = Font(bold=True, size=14, color="312E81")
    ws2.cell(2, 1, "")
    lines = (notes or []) + [
        "• Do NOT change the column header names in row 1.",
        "• Delete the sample rows (rows 2 onwards) before uploading your real questions.",
        "• Correct Answer must be exactly: A, B, C, or D (capital letter).",
        "• Difficulty must be exactly: low, medium, or high (lowercase).",
        "• role_tags and skill_tags: comma-separated values in one cell  e.g.  Java Developer, QA Engineer",
        "• Save the file as .xlsx before uploading.",
        "• Required columns are marked in INDIGO. Optional columns are in GRAY.",
    ]
    for i, line in enumerate(lines, 3):
        c = ws2.cell(i, 1, line)
        c.font = Font(size=11, color="1F2937")
        c.alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[1].height = 24

    return wb


# ══════════════════════════════════════════════════════
#  IMPORT HELPER
# ══════════════════════════════════════════════════════
async def _import_questions(file: UploadFile, db: Session, model_class, segment: str):
    filename = (file.filename or "").lower()
    content  = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        elif filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content))
        else:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file type. Please upload a .xlsx or .csv file."
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {str(e)}")

    if df.empty:
        raise HTTPException(status_code=400, detail="File has no data rows.")

    # Normalise column names: lowercase + underscores
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    # Drop completely empty rows
    df.dropna(how="all", inplace=True)

    created = 0
    errors  = []

    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed header + 1)
        try:
            rv = row.where(pd.notna(row), None).to_dict()
            _str = lambda k, default="": str(rv[k]).strip() if rv.get(k) not in (None, "") else default

            if segment == "seg3":
                scenario = _str("scenario_text")
                if not scenario:
                    errors.append({"row": row_num, "error": "scenario_text is required"})
                    continue
                obj = model_class(
                    scenario_text=scenario,
                    reference_answer=_str("reference_answer"),
                    difficulty=_str("difficulty", "medium").lower(),
                    role_tags=_parse_list(rv.get("role_tags")),
                    is_active=True,
                )
            else:
                question = _str("question_text")
                if not question:
                    errors.append({"row": row_num, "error": "question_text is required"})
                    continue
                for opt in ["option_a", "option_b", "option_c", "option_d"]:
                    if not _str(opt):
                        errors.append({"row": row_num, "error": f"{opt} is required"})
                        break
                else:
                    ans = _str("correct_answer", "A").upper()
                    if ans not in ("A", "B", "C", "D"):
                        errors.append({"row": row_num, "error": f"correct_answer must be A/B/C/D, got '{ans}'"})
                        continue
                    diff = _str("difficulty", "medium").lower()
                    if diff not in ("low", "medium", "high"):
                        diff = "medium"

                    obj = model_class(
                        question_text=question,
                        option_a=_str("option_a"),
                        option_b=_str("option_b"),
                        option_c=_str("option_c"),
                        option_d=_str("option_d"),
                        correct_answer=ans,
                        difficulty=diff,
                        category=_str("category") or None,
                        is_active=True,
                    )
                    if segment == "seg2":
                        obj.role_tags  = _parse_list(rv.get("role_tags"))
                        obj.skill_tags = _parse_list(rv.get("skill_tags"))

                    db.add(obj)
                    created += 1
                    continue
                continue  # skip on option error

            db.add(obj)
            created += 1

        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    db.commit()
    return {
        "imported": created,   # frontend reads this
        "created":  created,   # keep for compatibility
        "errors":   errors,
        "message":  f"{created} question(s) imported successfully."
                    + (f" {len(errors)} row(s) skipped — see errors." if errors else ""),
    }


def _parse_list(value) -> Optional[List[str]]:
    if value is None:
        return None
    if isinstance(value, list):
        return value
    s = str(value).strip()
    if not s:
        return None
    return [v.strip() for v in s.split(",") if v.strip()]


def _wb_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


# ══════════════════════════════════════════════════════
#  SERIALISERS
# ══════════════════════════════════════════════════════
def _q1_to_dict(q):
    return {
        "id": q.id, "question_text": q.question_text,
        "option_a": q.option_a, "option_b": q.option_b,
        "option_c": q.option_c, "option_d": q.option_d,
        "correct_answer": q.correct_answer, "difficulty": q.difficulty,
        "category": q.category, "experience_bracket_ids": q.experience_bracket_ids,
        "is_active": q.is_active, "usage_count": q.usage_count,
        "created_at": q.created_at,
    }

def _q2_to_dict(q):
    return {
        "id": q.id, "question_text": q.question_text,
        "option_a": q.option_a, "option_b": q.option_b,
        "option_c": q.option_c, "option_d": q.option_d,
        "correct_answer": q.correct_answer, "difficulty": q.difficulty,
        "category": q.category, "role_tags": q.role_tags,
        "skill_tags": q.skill_tags, "experience_bracket_ids": q.experience_bracket_ids,
        "is_active": q.is_active, "usage_count": q.usage_count,
        "created_at": q.created_at,
    }

def _q3_to_dict(q):
    return {
        "id": q.id, "scenario_text": q.scenario_text,
        "reference_answer": q.reference_answer, "difficulty": q.difficulty,
        "role_tags": q.role_tags, "is_active": q.is_active,
        "usage_count": q.usage_count, "created_at": q.created_at,
    }


# ══════════════════════════════════════════════════════
#  SEGMENT 1
# ══════════════════════════════════════════════════════
@router.get("/seg1/template")
def template_seg1(current_user=Depends(require_admin)):
    headers = [
        ("question_text",  True,  60),
        ("option_a",       True,  30),
        ("option_b",       True,  30),
        ("option_c",       True,  30),
        ("option_d",       True,  30),
        ("correct_answer", True,  14),
        ("difficulty",     True,  12),
        ("category",       False, 22),
    ]
    samples = [
        ["What does ACID stand for in database systems?",
         "Atomicity, Consistency, Isolation, Durability",
         "Association, Consistency, Integrity, Durability",
         "Atomicity, Concurrency, Isolation, Dependency",
         "Association, Concurrency, Integrity, Dependency",
         "A", "medium", "Databases"],
        ["Which protocol is used for secure data transfer over the web?",
         "HTTP", "FTP", "HTTPS", "SMTP",
         "C", "low", "Networking"],
    ]
    notes = [
        "Segment 1 — Knowledge MCQ (multiple choice, single correct answer)",
        "• category is optional but recommended for question organisation.",
        "",
    ]
    wb = _styled_workbook(headers, samples, notes, "Segment1_Questions")
    return _wb_response(wb, "TalentLens_Segment1_Template")


@router.get("/seg1")
def list_seg1(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    difficulty: Optional[str] = None, category: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db), current_user=Depends(require_admin),
):
    q = db.query(models.QuestionSeg1)
    if difficulty: q = q.filter(models.QuestionSeg1.difficulty == difficulty)
    if category:   q = q.filter(models.QuestionSeg1.category == category)
    if search:     q = q.filter(models.QuestionSeg1.question_text.ilike(f"%{search}%"))
    total = q.count()
    items = q.order_by(models.QuestionSeg1.id.desc()).offset((page-1)*page_size).limit(page_size).all()
    return {"items": [_q1_to_dict(i) for i in items], "total": total, "page": page}


@router.post("/seg1")
def create_seg1(payload: schemas.QuestionSeg1Create, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = models.QuestionSeg1(**payload.dict()); db.add(q); db.commit()
    return {"message": "Question created.", "id": q.id}


@router.put("/seg1/{qid}")
def update_seg1(qid: int, payload: schemas.QuestionSeg1Update, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = db.query(models.QuestionSeg1).filter_by(id=qid).first()
    if not q: raise HTTPException(404, "Not found")
    for f, v in payload.dict(exclude_none=True).items(): setattr(q, f, v)
    db.commit(); return {"message": "Updated."}


@router.delete("/seg1/purge")
def purge_seg1(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    count = db.query(models.QuestionSeg1).count()
    db.query(models.QuestionSeg1).delete()
    db.commit()
    return {"message": f"All {count} Segment 1 questions deleted.", "deleted": count}


@router.delete("/seg1/{qid}")
def delete_seg1(qid: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = db.query(models.QuestionSeg1).filter_by(id=qid).first()
    if not q: raise HTTPException(404, "Not found")
    q.is_active = False; db.commit()
    return {"message": "Deactivated."}


@router.post("/seg1/import")
async def import_seg1(file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    return await _import_questions(file, db, models.QuestionSeg1, "seg1")


@router.get("/seg1/export")
def export_seg1(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    items = db.query(models.QuestionSeg1).filter_by(is_active=True).all()
    if not items: raise HTTPException(404, "No active questions to export.")
    rows = [{"question_text": q.question_text, "option_a": q.option_a, "option_b": q.option_b,
             "option_c": q.option_c, "option_d": q.option_d, "correct_answer": q.correct_answer,
             "difficulty": q.difficulty, "category": q.category or ""} for q in items]
    wb = Workbook(); ws = wb.active; ws.title = "Segment1_Questions"
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows: ws.append([r[h] for h in headers])
    return _wb_response(wb, "talentlens_seg1_export")


# ══════════════════════════════════════════════════════
#  SEGMENT 2
# ══════════════════════════════════════════════════════
@router.get("/seg2/template")
def template_seg2(current_user=Depends(require_admin)):
    headers = [
        ("question_text",  True,  60),
        ("option_a",       True,  30),
        ("option_b",       True,  30),
        ("option_c",       True,  30),
        ("option_d",       True,  30),
        ("correct_answer", True,  14),
        ("difficulty",     True,  12),
        ("category",       False, 22),
        ("role_tags",      False, 30),
        ("skill_tags",     False, 30),
    ]
    samples = [
        ["A client reports intermittent connectivity issues. What is your first step?",
         "Restart all network equipment immediately",
         "Gather information and reproduce the issue",
         "Escalate to Tier 2 without investigation",
         "Ask the client to try again later",
         "B", "medium", "Troubleshooting",
         "L1 Support, Network Engineer",
         "Incident Management, Triage"],
        ["Which ITIL process handles unplanned interruptions to IT services?",
         "Change Management", "Incident Management", "Problem Management", "Service Request",
         "B", "low", "ITIL",
         "Service Desk, ITSM Analyst",
         "ITIL, Service Management"],
    ]
    notes = [
        "Segment 2 — Role Competency MCQ (role-specific questions)",
        "• role_tags: comma-separated roles this question applies to.",
        "• skill_tags: comma-separated skills being tested.",
        "",
    ]
    wb = _styled_workbook(headers, samples, notes, "Segment2_Questions")
    return _wb_response(wb, "TalentLens_Segment2_Template")


@router.get("/seg2")
def list_seg2(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    difficulty: Optional[str] = None, search: Optional[str] = None,
    db: Session = Depends(get_db), current_user=Depends(require_admin),
):
    q = db.query(models.QuestionSeg2)
    if difficulty: q = q.filter(models.QuestionSeg2.difficulty == difficulty)
    if search:     q = q.filter(models.QuestionSeg2.question_text.ilike(f"%{search}%"))
    total = q.count()
    items = q.order_by(models.QuestionSeg2.id.desc()).offset((page-1)*page_size).limit(page_size).all()
    return {"items": [_q2_to_dict(i) for i in items], "total": total, "page": page}


@router.post("/seg2")
def create_seg2(payload: schemas.QuestionSeg2Create, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = models.QuestionSeg2(**payload.dict()); db.add(q); db.commit()
    return {"message": "Question created.", "id": q.id}


@router.put("/seg2/{qid}")
def update_seg2(qid: int, payload: schemas.QuestionSeg2Update, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = db.query(models.QuestionSeg2).filter_by(id=qid).first()
    if not q: raise HTTPException(404, "Not found")
    for f, v in payload.dict(exclude_none=True).items(): setattr(q, f, v)
    db.commit(); return {"message": "Updated."}


@router.delete("/seg2/purge")
def purge_seg2(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    count = db.query(models.QuestionSeg2).count()
    db.query(models.QuestionSeg2).delete()
    db.commit()
    return {"message": f"All {count} Segment 2 questions deleted.", "deleted": count}


@router.delete("/seg2/{qid}")
def delete_seg2(qid: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = db.query(models.QuestionSeg2).filter_by(id=qid).first()
    if not q: raise HTTPException(404, "Not found")
    q.is_active = False; db.commit()
    return {"message": "Deactivated."}


@router.post("/seg2/import")
async def import_seg2(file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    return await _import_questions(file, db, models.QuestionSeg2, "seg2")


@router.get("/seg2/export")
def export_seg2(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    items = db.query(models.QuestionSeg2).filter_by(is_active=True).all()
    if not items: raise HTTPException(404, "No active questions to export.")
    rows = [{"question_text": q.question_text, "option_a": q.option_a, "option_b": q.option_b,
             "option_c": q.option_c, "option_d": q.option_d, "correct_answer": q.correct_answer,
             "difficulty": q.difficulty, "category": q.category or "",
             "role_tags": ",".join(q.role_tags or []), "skill_tags": ",".join(q.skill_tags or [])} for q in items]
    wb = Workbook(); ws = wb.active; ws.title = "Segment2_Questions"
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows: ws.append([r[h] for h in headers])
    return _wb_response(wb, "talentlens_seg2_export")


# ══════════════════════════════════════════════════════
#  SEGMENT 3
# ══════════════════════════════════════════════════════
@router.get("/seg3/template")
def template_seg3(current_user=Depends(require_admin)):
    headers = [
        ("scenario_text",    True,  80),
        ("reference_answer", True,  80),
        ("difficulty",       True,  12),
        ("role_tags",        False, 35),
    ]
    samples = [
        ["A major client's trading system has been down for 45 minutes during market hours. "
         "The Tier 1 team has exhausted their checklist. As the escalation engineer, describe your approach.",
         "Acknowledge urgency. Immediately review recent change logs and deployment activity. "
         "Check infrastructure health (DB, network, app servers). Form a bridge call with key stakeholders. "
         "Communicate status every 15 minutes. Engage vendor support if needed. Document RCA post-resolution.",
         "high",
         "L2 Support, Incident Manager, SRE"],
        ["You receive a task to onboard 50 new users to an enterprise banking application by end of day. "
         "Halfway through, you discover the bulk upload tool is broken. What do you do?",
         "Assess remaining count and time. Attempt manual onboarding for critical users first. "
         "Raise a P2 ticket for the broken tool. Communicate revised ETA to stakeholder. "
         "Explore scripted workaround if available. Escalate blockers proactively.",
         "medium",
         "Operations, Support Analyst"],
    ]
    notes = [
        "Segment 3 — Scenario Response (AI-evaluated open-ended answers)",
        "• scenario_text: The situation/problem presented to the candidate.",
        "• reference_answer: The ideal answer used as the AI evaluation benchmark.",
        "• The AI scores candidate responses against your reference_answer.",
        "",
    ]
    wb = _styled_workbook(headers, samples, notes, "Segment3_Scenarios")
    return _wb_response(wb, "TalentLens_Segment3_Template")


@router.get("/seg3")
def list_seg3(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    db: Session = Depends(get_db), current_user=Depends(require_admin),
):
    q = db.query(models.QuestionSeg3)
    if search: q = q.filter(models.QuestionSeg3.scenario_text.ilike(f"%{search}%"))
    total = q.count()
    items = q.order_by(models.QuestionSeg3.id.desc()).offset((page-1)*page_size).limit(page_size).all()
    return {"items": [_q3_to_dict(i) for i in items], "total": total, "page": page}


@router.post("/seg3")
def create_seg3(payload: schemas.QuestionSeg3Create, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = models.QuestionSeg3(**payload.dict()); db.add(q); db.commit()
    return {"message": "Question created.", "id": q.id}


@router.put("/seg3/{qid}")
def update_seg3(qid: int, payload: schemas.QuestionSeg3Update, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = db.query(models.QuestionSeg3).filter_by(id=qid).first()
    if not q: raise HTTPException(404, "Not found")
    for f, v in payload.dict(exclude_none=True).items(): setattr(q, f, v)
    db.commit(); return {"message": "Updated."}


@router.delete("/seg3/purge")
def purge_seg3(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    count = db.query(models.QuestionSeg3).count()
    db.query(models.QuestionSeg3).delete()
    db.commit()
    return {"message": f"All {count} Segment 3 questions deleted.", "deleted": count}


@router.delete("/seg3/{qid}")
def delete_seg3(qid: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    q = db.query(models.QuestionSeg3).filter_by(id=qid).first()
    if not q: raise HTTPException(404, "Not found")
    q.is_active = False; db.commit()
    return {"message": "Deactivated."}


@router.post("/seg3/import")
async def import_seg3(file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(require_admin)):
    return await _import_questions(file, db, models.QuestionSeg3, "seg3")


@router.get("/seg3/export")
def export_seg3(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    items = db.query(models.QuestionSeg3).filter_by(is_active=True).all()
    if not items: raise HTTPException(404, "No active questions to export.")
    rows = [{"scenario_text": q.scenario_text, "reference_answer": q.reference_answer,
             "difficulty": q.difficulty, "role_tags": ",".join(q.role_tags or [])} for q in items]
    wb = Workbook(); ws = wb.active; ws.title = "Segment3_Scenarios"
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows: ws.append([r[h] for h in headers])
    return _wb_response(wb, "talentlens_seg3_export")
